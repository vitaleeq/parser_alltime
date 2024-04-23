import string
import parsing_functions as pf
from openpyxl import load_workbook


def use_excel_file(filepath):
    sheet_name = input('Input sheet name: ')
    lst = input("Input vendor_code's column letter, first_row and last_row(e.g. B, 10, 73): ").split(', ')
    vendor_codes_col = [int(elem) if elem.isdigit() else string.ascii_uppercase.find(elem.upper()) for elem in lst]
    # Можно добавить и использовать тут функцию, которая аналогично select_required_cols автоматически будет искать
    # и использовать столбец с названием "Артикул"
    # Here we got smth like this: [1{0-26}, 10{1-inf}, 73{1-inf}]

    # load file, activate sheet in it, select targeted cell's
    wb = load_workbook(filename=filepath)
    wb_sheet = wb.active if sheet_name == '' else wb[sheet_name]
    required_cols = select_required_cols(vendor_codes_col[1] - 1, wb_sheet)
    # Here we got coords of required cols
    print(f'{required_cols=}')
    print('OOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOO')
    try:
        pf.create_driver()
        for line in range(vendor_codes_col[1], vendor_codes_col[2]+1):
            vendor_code = get_vendor_code(wb_sheet, line, vendor_codes_col[0])
            print(f'{wb_sheet[line][vendor_codes_col[0]]}: {vendor_code}')
            link = pf.get_product_link(vendor_code)
            print(f'{vendor_code}: {link}')
            print('___________________________________')
            properties = pf.get_properties(link)
            print(f'required_properties = {properties}')
            print(f'{required_cols=}')
            print('***********************************')
            input_properties(wb_sheet, line, properties, required_cols)
    except KeyboardInterrupt:
        pf.quit_driver()
    pf.quit_driver()

    '''
    # create dictionary with keys==cell's coords, values==cell's values
    cells_with_serial_numbers = ft.create_cells_dictionary(wb_sheet, cell_names)

    # change serial numbers for each cell using randomiser, rewrite new serial numbers in file and save it
    for key, values in cells_with_serial_numbers.items():
        wb_sheet[key] = ', '.join([ft.random_last_symbols(value) for value in values])
    '''

    wb.save(f"new_{create_new_filename(filepath)}")


def create_new_filename(filepath):
    return filepath if '/' not in filepath else filepath[filepath.rfind('/')+1:]


def select_required_cols(attrs_row_number, wb_sheet):
    search_dict = {'body': 'Корпус',
                   'back_cover': 'Задняя крышка',
                   'mechanism_type': 'Тип механизма',
                   'bracelet': 'Тип браслета',
                   'glass': 'Стекло',
                   'additional_functions': 'Дополнительные функции',
                   'insertions': 'Вставки'}
    required_cols_dict = {}
    for elem in search_dict:
        for row in wb_sheet.iter_rows(min_row=attrs_row_number, max_row=attrs_row_number, min_col=wb_sheet.min_column, max_col=wb_sheet.max_column):
            for cell in row:
                if cell.value is None:
                    continue
                if search_dict[elem].strip().lower() in cell.value.strip().lower():
                    required_cols_dict[elem] = cell.column
    return required_cols_dict


def get_vendor_code(wb_sheet, row, col):
    vendor_code = wb_sheet[row][col].value
    return vendor_code


def input_properties(wb_sheet, line, properties, required_cols):
    for key in properties:
        #print(f'{key=}')
        #print(f'{line=}')
        #print(f'{required_cols[key]=}')
        cell = wb_sheet[line][required_cols[key]-1]
        #print(cell)
        cell.value = properties[key]
        print(f'{cell}: {cell.value}')
        print('\n')
